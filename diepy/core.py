import csv
import codecs
from ConfigParser import SafeConfigParser
from datetime import datetime
import gzip
import logging
import os
from os import path
import re

from dateutil.parser import parse
import openpyxl
import sqlalchemy

logger = logging.getLogger(__name__)


def is_csv(filepath):
    regex = re.compile(".*(.csv|.tab|.tsv|.txt)(.gz|.zip)?$")
    return regex.match(filepath)


def is_excel(filepath):
    regex = re.compile(".*(.xlsx|.xls)(.gz|.zip)?(\$.*)?$")
    return regex.match(filepath)


def parse_dbpath(dbpath):
    parts = dbpath.split('.')

    server = None
    database = None
    schema = None
    table = None

    if len(parts) == 1:
        server = parts[0]
    elif len(parts) == 2:
        server = parts[0]
        database = parts[1]
    elif len(parts) == 3:
        server = parts[0]
        database = parts[1] or None
        schema = parts[2]
    elif len(parts) == 4:
        server = parts[0]
        database = parts[1] or None
        schema = parts[2] or None
        table = parts[3]

    return server, database, schema, table

def import_files(import_path, server, database=None, schema=None, table=None, delimiter=None, config=None):
    """Import file(s) into database.
    
    Args:
        server (string): the server to connect to. Should be in config file.
        database: the database on the server to connect to.
        schema: the schema of the table.
        table: the name of the table
        delimiter: the delimiter to use when parsing the text file. default is comma
        import_paths: either a single file or a directory of files.
        config: path to a specific configuration file to use.
    
    Returns:
        None
    
    """
    db = Database(server, database, config)

    if path.isfile(import_path):
        db.import_file(import_path, table, schema, delimiter=delimiter)
        return

    for fpath in [path.join(import_path, p) for p in os.listdir(import_path)]:
        if not fpath.endswith('.csv'):
            continue
        logger.info("Importing: %s" % fpath)
        db.import_file(fpath, None, schema, delimiter=delimiter)


def export_table(server, database, schema, table, export_path):
    db = Database(server, database)
    db.export_table(table, export_path)


class Database(object):
    def __init__(self, server, database=None, config=None):
        self.metadata = sqlalchemy.MetaData()
        self.engine = None

        parser = SafeConfigParser()

        if config:
            logger.info('Parsing config file %s' % config)
            parser.read(config)
        elif path.exists('diepy.ini'):
            logger.info('Parsing config file %s' % path.abspath('diepy.ini'))
            parser.read('diepy.ini')
        elif path.exists(path.expanduser("~/diepy.ini")):
            logger.info('Parsing config file %s' % path.abspath(path.expanduser('~/diepy.ini')))
            parser.read(path.abspath(path.expanduser('~/diepy.ini')))
        else:
            raise Exception("No configuration file found!")

        cstring = parser.get('servers', server)

        if database:
            cstring = cstring.rstrip('/') + '/' + database

        logger.info('Connecting to database: %s' % cstring)
        self.engine = sqlalchemy.create_engine(cstring)
        self.metadata.bind = self.engine

    def import_file(self, filepath, table_name=None, schema=None, delimiter=',', truncate=False):
        """Import a file into the database.

        Args:
            filepath (str): the path to the file to import.
            table_name (str): the particular table to import into. defaults to be based on the filename sans extension
            schema (str): the particular schema to import into. defaults to database/user default
            delimiter (char): the delimiter to use. default to comma (,)

        Returns:
            int: the number of rows imported

        """

        try:
            if is_excel(filepath):
                self.import_excel(filepath, table_name, schema, truncate)
            else:
                if not table_name:
                    (table_name, ext) = path.splitext(path.basename(filepath))

                if not self.table_exists(table_name, schema):
                    columns = generate_schema_from_csv(filepath, delimiter)
                    self.create_table(table_name, columns, schema=schema)

                table = sqlalchemy.Table(table_name, self.metadata, autoload=True, schema=schema)
                if truncate:
                    self.engine.execute(table.delete())

                if filepath.endswith('.xlsx'):
                    rows = self.store_xlsx(filepath, table)
                else:
                    rows = self.store_data(filepath, table, delimiter)
                return rows

        except:
            logger.exception("Had some trouble storing %s" % filepath)

    def import_excel(self, filepath, table_name, schema=None, truncate=False):
        if '$' in filepath:
            f, sheet = filepath.split('$')
        else:
            f = filepath
            sheet = None

        wb = openpyxl.load_workbook(filename=f, use_iterators=True)

        if sheet:
            self.import_worksheet(wb, sheet, table_name, schema, truncate)
        else:
            for sheet in wb.get_sheet_names():
                self.import_worksheet(wb, sheet, table_name, schema, truncate)

    def import_worksheet(self, wb, sheet, table_name, schema, truncate=False):
        logger.info("Importing worksheet '{}'...".format(sheet))
        table_name = table_name or sheet
        ws = wb.get_sheet_by_name(sheet)

        if not self.table_exists(table_name, schema):
            columns = generate_schema_from_excel(ws)
            self.create_table(table_name, columns, schema=schema)

        table = sqlalchemy.Table(table_name, self.metadata, autoload=True, schema=schema)
        if truncate:
            self.engine.execute(table.delete())

        rows = self.store_xlsx(ws, table)

        return rows

    def table_exists(self, table_name, schema=None):
        """Determines if a table already exists in the database

        Args:
            table_name (str): the name of the table to check for
            schema (str): optional - the schema to search in. defaults to database/user default schema.

        Returns:
            bool: whether or not the table exists

        """

        exists = self.engine.dialect.has_table(self.engine.connect(), table_name, schema=schema)
        if not exists:
            logger.warn("Table '%s' does not exist." % table_name)
        return exists

    def create_table(self, table_name, columns, schema=None):
        table = sqlalchemy.Table(table_name, self.metadata, schema=schema)
        if not columns:
            logger.warn("No columns found")
            return

        for col in columns:
            table.append_column(col.emit())

        table.create(self.engine)

    def create_table_from_csv(self, table_name, filepath, delimiter, schema=None):
        logger.info("Creating table for '%s' ..." % filepath)
        table = sqlalchemy.Table(table_name, self.metadata, schema=schema)
        if is_csv(filepath):
            columns = generate_schema_from_csv(filepath, delimiter)
        elif is_excel(filepath):
            columns = generate_schema_from_excel(filepath)
        else:
            raise Exception("Unknown file type: %s" % filepath)

        if not columns:
            logger.warn("No columns found")
            return

        for col in columns:
            table.append_column(col.emit())

        table.create(self.engine)

    def store_data(self, filepath, table, delimiter=','):
        logger.info("Storing records from %s in %s" % (filepath, table.name))
        cn = self.engine.connect()
        infile = open(filepath, 'rbU')
        dr = csv.DictReader(infile, delimiter=delimiter)
        rows = 0
        batch = []
        for row in dr:
            rows += 1
            record = {k.replace(codecs.BOM_UTF8, ''): cast_data(k, v, table) for k, v in row.items()}
            batch.append(record)
            if rows % 1000 == 0:
                cn.execute(table.insert(), batch)
                logger.info("Imported %s records..." % rows)
                batch = []

        if batch:
            cn.execute(table.insert(), batch)

        logger.info("Stored %s records from %s in %s" % (rows, filepath, table.name))
        return rows

    def store_xlsx(self, ws, table, sheet=None):
        cn = self.engine.connect()

        rows = -1
        header = None
        batch = []
        for row in ws.iter_rows():
            rows += 1
            if rows == 0:
                header = [c.internal_value for c in row]
                continue

            values = [c.internal_value for c in row]
            zipped = zip(header, values)
            record = dict(zipped)
            # record = {k: cast_data(k, v, table) for k, v in zipped}
            batch.append(record)

            if rows % 1000 == 0:
                cn.execute(table.insert(), batch)
                logger.info("Imported %s records..." % rows)
                batch = []

        if batch:
            cn.execute(table.insert(), batch)

        if rows == -1:
            logger.warn("No data found.")

    def export_table(self, table, filename, schema=None, unix=False, zip=False):
        mytable = sqlalchemy.Table(table, self.metadata, autoload=True, schema=schema or None)
        db_connection = self.engine.connect()

        select = sqlalchemy.sql.select([mytable])
        results = db_connection.execute(select)
        logger.info(filename)
        if is_excel(filename):
            self.write_xlsx(filename, table, results)
        else:
            self.write_csv(filename, results, unix, zip)

    def write_csv(self, filename, data, unix=False, windows=False, zip=False):
        if zip:
            if not filename.endswith('.gz'):
                filename += '.gz'
            f = gzip.open(filename, 'wb')
        else:
            f = open(filename, 'wb')

        if unix:
            lineterminator = '\n'
        elif windows:
            lineterminator = '\r\n'
        else:
            lineterminator = os.linesep

        if filename.endswith('.tab') or filename.endswith('.tsv'):
            delimiter = '\t'
        else:
            delimiter = ','

        try:
            fieldnames = data.keys()
            writer = csv.DictWriter(f,
                    lineterminator=lineterminator,
                    delimiter=delimiter,
                    fieldnames=fieldnames)
            headers = dict((n, unicode(n)) for n in fieldnames)
            writer.writerow(headers)
            records = 0
            for row in data:
                cleaned = {k: self._cleanbool(v) for k, v in row.items()}
                writer.writerow(cleaned)
                records += 1

        finally:
            f.close()

        logger.info("Wrote %s records to %s" % (records, filename))

    def write_xlsx(self, filename, tablename, data):
        if path.isfile(filename):
            wb = openpyxl.load_workbook(filename)
            ws = wb.get_sheet_by_name(tablename)
            if ws:
                wb.remove_sheet(ws)
            ws = wb.create_sheet()
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
        
        ws.title = tablename
        
        for c, title in enumerate(data.keys()):
            ws.cell(row=0, column=c).value = title
        
        for r, row in enumerate(data):
            for c, value in enumerate(row):
                ws.cell(row=r + 1, column=c).value = value
        
        wb.save(filename=filename)

    @staticmethod
    def _cleanbool(value):
        if value is None:
            return value

        if type(value) is bool and value:
            return 1
        if type(value) is bool and not value:
            return 0
        if type(value) is datetime:
            return value.isoformat()
        return unicode(value)


def cast_data(k, v, table):
    k = k.replace(codecs.BOM_UTF8, '')
    if v is None or v == '':
        return None

    # logger.debug('Attempting to cast %s as %s ...' % (v, table.c[k].type))
    if isinstance(table.c[k].type, sqlalchemy.types.DATETIME) or isinstance(table.c[k].type, sqlalchemy.types.DATE):
        v = parse(v)
        return v

    if isinstance(table.c[k].type, sqlalchemy.types.TIME):
        dt = parse(v)
        v = dt.time()
        return v

    return unicode(v, 'utf-8')


def cast_datetime(v):
    logger.info('Attempting to cast %s as datetime...' % v)
    try:
        return datetime.strptime(v, "%m/%d/%y")
    except ValueError:
        pass

    try:
        return datetime.strptime(v, "%Y-%m-%d")
    except ValueError:
        pass

    return v


def generate_schema_from_excel(ws, sheet=None, sample_size=20000):
    """Generates a table DDL statement based on the file"""
    logger.info("Generating schema for '%s'" % ws)

    columns = []
    samples = -1
    unnamed = 0
    for row in ws.iter_rows():
        samples += 1
        if samples == 0:
            for h in row:
                if h.internal_value is None:
                    unnamed += 1
                columns.append(ColumnDef(h.internal_value or "unnamed%s" % unnamed))
            continue

        for i, c in enumerate(row):
            columns[i].sample_value(c.internal_value)
        if samples == sample_size:
            break

    return columns


def generate_schema_from_csv(filepath, delimiter=',', sample_size=20000):
    """Generates a table DDL statement based on the file"""
    logger.info("Generating schema for '%s'" % filepath)
    infile = open(filepath, 'rbU')
    dr = csv.reader(infile, delimiter=delimiter)
    
    columns = []
    samples = -1
    unnamed = 0
    for row in dr:
        samples += 1
        if samples == 0:
            for h in row:
                h = h.strip().replace(codecs.BOM_UTF8, '')
                if h is None or h == '':
                    unnamed += 1
                    h = "unnamed%s" % unnamed
                columns.append(ColumnDef(h))
            continue
        for i, c in enumerate(row):
            columns[i].sample_value(c)
        if samples == sample_size:
            break

    return columns


class ColumnDef(object):
    """Defines the schema for a table column"""

    def __init__(self, name=''):
        self.name = name or 'unnamed'
        self.nullable = False
        self.type = ''
        self.length = 0
        self.max_value = 0
        self.min_value = 0

    def sample_value(self, value):
        """Samples a value to determine the datatype for the column"""

        if value == '' or value is None:
            self.nullable = True
            return

        self._determine_type(value)

        if self.type == 'text' and len(str(value)) > self.length:
            self.length = len(str(value))

        if self.type == 'int':
            if int(value) < self.min_value:
                self.min_value = int(value)
            
            if int(value) > self.max_value:
                self.max_value = int(value)

    def _determine_type(self, value):
        if self.type == 'date' and not is_date(value):
            self.type = 'text'
        if self.type == 'float' and not is_float(value):
            self.type = 'text'
        if self.type == 'int' and not is_int(value):
            self.type = 'text'

        if self.type == '':
            if is_int(value):
                self.type = 'int'
            elif is_float(value):
                self.type = 'float'
            elif is_date(value):
                self.type = 'date'
            elif is_time(value):
                self.type = 'time'
            elif is_datetime(value):
                self.type = 'datetime'
            else:
                self.type = 'text'

    def emit(self):
        """Prints the DDL for defining the column"""
        # logger.debug("emitting '%s' as '%s'. Nullable: %s" % (self.name, self.type, self.nullable))
        if self.type == '':
            self.type = 'text'

        if self.type == 'int' and self.max_value == 1 and self.min_value == 0:
            return sqlalchemy.Column(self.name, sqlalchemy.types.SMALLINT, nullable=self.nullable)
        elif self.type == 'int' and self.max_value >= 32768:
            return sqlalchemy.Column(self.name, sqlalchemy.types.INT, nullable=self.nullable)
        elif self.type == 'int':
            return sqlalchemy.Column(self.name, sqlalchemy.types.SMALLINT, nullable=self.nullable)
        elif self.type == 'float':
            return sqlalchemy.Column(self.name, sqlalchemy.types.FLOAT, nullable=self.nullable)
        elif self.type == 'datetime':
            return sqlalchemy.Column(self.name, sqlalchemy.types.DATETIME, nullable=self.nullable)
        elif self.type == 'date':
            return sqlalchemy.Column(self.name, sqlalchemy.types.DATE, nullable=self.nullable)
        elif self.type == 'time':
            return sqlalchemy.Column(self.name, sqlalchemy.types.TIME, nullable=self.nullable)
        elif self.type == 'text':
            if self.length < 50:
                return sqlalchemy.Column(self.name, sqlalchemy.Unicode(50), nullable=self.nullable)
            elif self.length < 100:
                return sqlalchemy.Column(self.name, sqlalchemy.Unicode(100), nullable=self.nullable)
            elif self.length < 200:
                return sqlalchemy.Column(self.name, sqlalchemy.Unicode(200), nullable=self.nullable)
            elif self.length < 500:
                return sqlalchemy.Column(self.name, sqlalchemy.Unicode(500), nullable=self.nullable)
            elif self.length < 1000:
                return sqlalchemy.Column(self.name, sqlalchemy.Unicode(1000), nullable=self.nullable)
            elif self.length < 4000:
                return sqlalchemy.Column(self.name, sqlalchemy.Unicode(4000), nullable=self.nullable)
            else:
                return sqlalchemy.Column(self.name, sqlalchemy.types.UnicodeText, nullable=self.nullable)


def is_int(s):
    if isinstance(s, (int, long)):
        return True

    # s = str(s)
    # if s.endswith('.0'):
    #     s = s.split('.')[0]
    # if s.startswith('0'):
    #     return False
    try:
        int(s)
        return True
    except ValueError:
        # logger.debug("Value is not an INT: %s" % s)
        return False


def is_float(s):
    s = str(s)
    try:
        float(s)
        # logger.debug("IS A FLOAT: %s" % s)
        return True
    except ValueError:
        # logger.debug("Value is not a FLOAT: %s" % s)
        return False


def is_time(s):

    try:
        # parse with two different default dates
        d1 = datetime(2000, 01, 01, 12, 34, 56, 123456)
        d2 = datetime(2007, 10, 20, 14, 32, 12, 654321)
        v1 = parse(str(s), default=d1)
        v2 = parse(str(s), default=d2)

        # if the year/month/day end up matching both default dates, then we got time only
        if d1.timetuple()[:3] == v1.timetuple()[:3] and d2.timetuple()[:3] == v2.timetuple()[:3]:
            return True
    except:
        pass

    # logger.debug("Value is not a TIME: %s" % s)
    return False


def is_date(s):
    try:
        # parse with two different default dates
        d1 = datetime(2000, 01, 01, 12, 34, 56)
        d2 = datetime(2007, 10, 20, 14, 32, 12)
        v1 = parse(str(s), default=d1)
        v2 = parse(str(s), default=d2)

        # if the hour, minute, second, microseconds end up matching both default dates, then we got date only
        if d1.timetuple()[3:6] == v1.timetuple()[3:6] and d2.timetuple()[3:6] == v2.timetuple()[3:6]:
            return True
        if parse(s).timetuple()[3:6] == (0, 0, 0):
            return True
    except:
        pass

    # logger.debug("Value is not a DATE: %s" % s)
    return False


def is_datetime(s):
    try:
        # parse with two different default dates
        d1 = datetime(2000, 01, 01, 12, 34, 56, 123456)
        d2 = datetime(2007, 10, 20, 14, 32, 12, 654321)
        v1 = parse(str(s), default=d1)
        v2 = parse(str(s), default=d2)

        # if if doesn't match either, then we've got a date time
        if d1 != v1 and d2 != v2:
            logger.debug("IS A DATETIME: %s" % s)
            return True
    except:
        pass

    # logger.debug("Value is not a DATETIME: %s" % s)
    return False
